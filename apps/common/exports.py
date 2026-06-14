try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_excel(data, filename, headers=None):
    """
    Génère un fichier Excel à partir d'une liste de dictionnaires.
    """
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    if headers:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(data, 2):
        if isinstance(item, dict):
            for col_idx, key in enumerate(headers or item.keys(), 1):
                ws.cell(row=row_idx, column=col_idx, value=str(item.get(key, '')))
        elif isinstance(item, (list, tuple)):
            for col_idx, val in enumerate(item, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
